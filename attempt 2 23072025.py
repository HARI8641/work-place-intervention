import os
import sys
import hashlib
import datetime
import subprocess
import tkinter as tk
from tkinter import messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageDraw, ImageFont, ImageTk
from tkcalendar import Calendar
import openpyxl
import qrcode
import re
import shutil
import platform

# --- CONFIG PATHS ---
BASE_DIR = os.path.join(os.path.expanduser("~"), "Documents", "id_gen_admin")
EXCEL_FILE = os.path.join(BASE_DIR, "data_base", "patient_data.xlsx")
ID_OUTPUT_DIR = os.path.join(BASE_DIR, "gen_id")
LOGO_FILE = os.path.join(BASE_DIR, "logo", "logo.png")
LICENSE_DIR = os.path.join(BASE_DIR, "logo", "license")
CRED_FILE = os.path.join(LICENSE_DIR, "cred.txt")
ADMIN_FILE = os.path.join(LICENSE_DIR, "admin.txt")
START_DATE_FILE = os.path.join(LICENSE_DIR, "start_date.txt")
PICTURES_DIR = os.path.join(os.path.expanduser("~"), "Pictures")
PICTURES_SUBDIR = os.path.join(PICTURES_DIR, "GKNMH_ID_Generator")
PICTURES_EXCEL = os.path.join(PICTURES_SUBDIR, "patient_data_pictures.xlsx")

# --- FORM WIDGETS ---
name_entry = None
dob_entry = None
gender_combobox = None
care_of_entry = None
phone_entry = None
calendar_widget = None
age_var = None

FOCUS_BG = "#e6f0ff"
ERROR_BORDER_COLOR = "#ff4d4d"
NORMAL_BORDER_COLOR = "#cccccc"

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)
        widget.bind("<Motion>", self.move_tip)

    def show_tip(self, event=None):
        if self.tipwindow or not self.text: return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 10
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("Segoe UI", "10"), wraplength=300)
        label.pack(ipadx=5, ipady=3)

    def move_tip(self, event):
        if self.tipwindow:
            x = event.x_root + 20
            y = event.y_root + 10
            self.tipwindow.wm_geometry(f"+{x}+{y}")

    def hide_tip(self, event=None):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None

class InteractiveEntry(tk.Entry):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.default_bg = self["bg"]
        self.error = False
        self.configure(relief="solid", bd=1, highlightthickness=0, highlightbackground=NORMAL_BORDER_COLOR)
        self.bind("<FocusIn>", self.on_focus_in)
        self.bind("<FocusOut>", self.on_focus_out)

    def on_focus_in(self, event):
        if not self.error:
            self.configure(background=FOCUS_BG, highlightbackground="#0078d7", highlightcolor="#0078d7", bd=2)

    def on_focus_out(self, event):
        if not self.error:
            self.configure(background="white", highlightbackground=NORMAL_BORDER_COLOR, bd=1)

    def mark_error(self, is_error):
        self.error = is_error
        if is_error:
            self.configure(background="#fff0f0", highlightbackground=ERROR_BORDER_COLOR, bd=2)
        else:
            self.configure(background="white", highlightbackground=NORMAL_BORDER_COLOR, bd=1)

def log_user_status(user_type, status):
    os.makedirs(LICENSE_DIR, exist_ok=True)
    log_file = os.path.join(LICENSE_DIR, "user_login_log.txt")
    with open(log_file, "a") as f:
        f.write(f"{datetime.datetime.now().isoformat()} | {user_type} login status: {status}\n")

def password_dialog(title, prompt, require_confirm=False, parent=None):
    pw = [None]
    dlg = tk.Toplevel(parent)
    dlg.title(title)
    dlg.geometry("400x180" if not require_confirm else "400x250")
    dlg.resizable(False, False)
    dlg.transient(parent)
    dlg.grab_set()
    entry_var = tk.StringVar()
    confirm_var = tk.StringVar()
    show_var = tk.IntVar()
    def toggle_show():
        show = "" if show_var.get() else "*"
        entry.config(show=show)
        if require_confirm:
            confirm_entry.config(show=show)
    def on_ok():
        pwd = entry_var.get()
        if require_confirm:
            confirm_pwd = confirm_var.get()
            if pwd != confirm_pwd:
                messagebox.showerror("Mismatch", "Passwords do not match.", parent=dlg)
                return
        pw[0] = pwd
        dlg.destroy()
    def on_cancel():
        dlg.destroy()
    tk.Label(dlg, text=prompt).pack(pady=(10, 5))
    entry = tk.Entry(dlg, textvariable=entry_var, show="*")
    entry.pack(pady=5, padx=20, fill="x")
    if require_confirm:
        tk.Label(dlg, text="Confirm Password:").pack(pady=(10, 2))
        confirm_entry = tk.Entry(dlg, textvariable=confirm_var, show="*")
        confirm_entry.pack(pady=2, padx=20, fill="x")
    chk = tk.Checkbutton(dlg, text="Show Password", variable=show_var, command=toggle_show)
    chk.pack(pady=5)
    btn_frame = tk.Frame(dlg)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="OK", width=10, command=on_ok).grid(row=0, column=0, padx=5)
    tk.Button(btn_frame, text="Cancel", width=10, command=on_cancel).grid(row=0, column=1, padx=5)
    entry.focus_set()
    dlg.wait_window()
    return pw[0]

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def is_strong_password(pw):
    return (len(pw) >= 8 and re.search(r"[A-Z]", pw) and re.search(r"[a-z]", pw)
            and re.search(r"[0-9]", pw) and re.search(r"[!@#$%^&*(),.?\":{}|<>]", pw))

def setup_dirs_and_files():
    os.makedirs(os.path.join(BASE_DIR, "data_base"), exist_ok=True)
    os.makedirs(ID_OUTPUT_DIR, exist_ok=True)
    os.makedirs(LICENSE_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(LOGO_FILE), exist_ok=True)
    os.makedirs(PICTURES_SUBDIR, exist_ok=True)
    if not os.path.exists(PICTURES_EXCEL):
        wb_pic = openpyxl.Workbook()
        ws_pic = wb_pic.active
        ws_pic.title = "Patient Data Pictures"
        ws_pic.append(["Patient ID", "Name", "DOB", "Age", "Gender", "Care Of", "Phone", "Registration Date", "Timestamp"])
        wb_pic.save(PICTURES_EXCEL)
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Patient ID", "Name", "DOB", "Age", "Gender", "Care Of", "Phone", "QR Path", "Reg Date", "Timestamp"])
        wb.save(EXCEL_FILE)
    if not os.path.exists(LOGO_FILE):
        Image.new("RGB", (600, 200), "gray").save(LOGO_FILE)
    if not os.path.exists(ADMIN_FILE):
        with open(ADMIN_FILE, "w") as f:
            f.write(hash_password("Admin@123"))
    if not os.path.exists(START_DATE_FILE):
        with open(START_DATE_FILE, "w") as f:
            f.write(datetime.datetime.today().strftime("%d-%m-%Y"))
    if os.name == 'nt':
        subprocess.call(["attrib", "+h", BASE_DIR])

def read_credentials():
    if not os.path.exists(CRED_FILE):
        return None, None, 0
    with open(CRED_FILE) as f:
        lines = f.readlines()
    if len(lines) < 3:
        return None, None, 0
    return lines[0].strip(), lines[1].strip(), int(lines[2].strip())

def is_user_password_expired():
    _, timestamp, _ = read_credentials()
    if not timestamp:
        return True
    try:
        last_set = datetime.datetime.fromisoformat(timestamp)
        return (datetime.datetime.now() - last_set).days > 90
    except:
        return True

def save_credentials(user_hash, timestamp, attempts=0):
    with open(CRED_FILE, "w") as f:
        f.write(f"{user_hash}\n{timestamp}\n{attempts}")

def read_admin_hash():
    if not os.path.exists(ADMIN_FILE): return ""
    with open(ADMIN_FILE) as f: return f.read().strip()

def write_admin_hash(new_hash):
    with open(ADMIN_FILE, "w") as f: f.write(new_hash)

def set_new_user_password(require_admin_auth=False, parent=None):
    if require_admin_auth:
        admin_pw = password_dialog("Admin Confirmation", "Enter Admin Password:", parent=parent)
        if admin_pw is None or hash_password(admin_pw) != read_admin_hash():
            messagebox.showerror("Access Denied", "Admin password invalid. Cannot proceed.", parent=parent)
            return False
    while True:
        pw = password_dialog("Set User Password", "Enter new password:", require_confirm=True, parent=parent)
        if pw is None: return False
        if not is_strong_password(pw):
            messagebox.showerror("Weak Password",
                "Password must be at least 8 characters long, include uppercase, lowercase, number, and symbol.",
                parent=parent)
            continue
        save_credentials(hash_password(pw), datetime.datetime.now().isoformat())
        messagebox.showinfo("Success", "User password has been set.", parent=parent)
        return True

def change_password(which, parent=None):
    if which == "User":
        return set_new_user_password(require_admin_auth=True, parent=parent)
    while True:
        pw = password_dialog(f"Set {which} Password", f"Enter new {which.lower()} password:", require_confirm=True, parent=parent)
        if pw is None: return False
        if not is_strong_password(pw):
            messagebox.showerror("Weak Password", "Password must include uppercase, lowercase, number, and symbol.", parent=parent)
            continue
        if which == "Admin":
            write_admin_hash(hash_password(pw))
        else:
            save_credentials(hash_password(pw), datetime.datetime.now().isoformat())
        messagebox.showinfo("Success", f"{which} password has been changed.", parent=parent)
        return True

def user_login_check(parent=None):
    stored_hash, _, attempts = read_credentials()
    if stored_hash is None or is_user_password_expired():
        if stored_hash is None:
            messagebox.showinfo("Setup User Password", "No user password found. Please set one now.", parent=parent)
        else:
            messagebox.showinfo("Password Expired", "User password expired. Please set a new one.", parent=parent)
        log_user_status("User", "Password setup or renewal required")
        if not set_new_user_password(require_admin_auth=True, parent=parent):
            log_user_status("User", "Password setup cancelled")
            return False
        log_user_status("User", "Password set or renewed")
        return True
    for i in range(3 - attempts):
        pw = password_dialog("User Login", "Enter user password:", parent=parent)
        if pw is None:
            log_user_status("User", "Login cancelled")
            return False
        if hash_password(pw) == stored_hash:
            save_credentials(stored_hash, datetime.datetime.now().isoformat(), 0)
            log_user_status("User", "Login successful")
            return True
        else:
            attempts += 1
            save_credentials(stored_hash, datetime.datetime.now().isoformat(), attempts)
            messagebox.showerror("Error", f"Invalid password. Attempts left: {3 - attempts}", parent=parent)
    admin_pw = password_dialog("Admin Override", "3 failed user login attempts. Enter admin password:", parent=parent)
    if admin_pw is None:
        log_user_status("User", "Admin override cancelled")
        return False
    if hash_password(admin_pw) == read_admin_hash():
        messagebox.showinfo("Admin Override", "Admin override granted. Please set a new user password.", parent=parent)
        log_user_status("User", "Admin override granted, resetting user password")
        if not set_new_user_password(parent=parent):
            log_user_status("User", "User password setup cancelled after admin override")
            return False
        return True
    else:
        messagebox.showerror("Access Denied", "Invalid admin password. Application locked.", parent=parent)
        log_user_status("User", "Admin override failed, application locked")
        return False

def admin_login_check(parent=None):
    for _ in range(3):
        pw = password_dialog("Admin Login", "Enter admin password:", parent=parent)
        if pw is None:
            log_user_status("Admin", "Login cancelled")
            return False
        if hash_password(pw) == read_admin_hash():
            log_user_status("Admin", "Login successful")
            return True
        else:
            messagebox.showerror("Invalid Password", "Incorrect admin password.", parent=parent)
    log_user_status("Admin", "Failed login attempts exceeded")
    return False

def admin_password_management_gui(parent=None):
    admin_window = tk.Toplevel(parent)
    admin_window.title("Admin Console - Management")
    admin_window.geometry("430x300")
    admin_window.configure(bg="#f6f8fa")
    admin_window.resizable(False, False)

    title = tk.Label(admin_window, text="Admin Console", font=("Segoe UI", 16, "bold"), bg="#0078d7", fg="white", pady=10)
    title.pack(fill="x")

    # Add icon img or logo if LOGO_FILE exists
    if os.path.exists(LOGO_FILE):
        try:
            logo_img = Image.open(LOGO_FILE).resize((160, 54))
            logo_img_tk = ImageTk.PhotoImage(logo_img)
            logo_label = tk.Label(admin_window, image=logo_img_tk, bg="#f6f8fa")
            logo_label.pack(pady=(10,2))
            logo_label.image = logo_img_tk
        except Exception:
            pass

    btn_user_pw = tk.Button(admin_window, text="Change User Password", width=30, command=lambda: change_password("User", parent=admin_window),
                            bg="#008080", fg="white", font=("Segoe UI", 12, "bold"), pady=8, relief="raised", cursor="hand2")
    btn_user_pw.pack(pady=12)

    btn_admin_pw = tk.Button(admin_window, text="Change Admin Password", width=30, command=lambda: change_password("Admin", parent=admin_window),
                             bg="#8700d8", fg="white", font=("Segoe UI", 12, "bold"), pady=8, relief="raised", cursor="hand2")
    btn_admin_pw.pack(pady=4)

    btn_close = tk.Button(admin_window, text="Close Console", command=admin_window.destroy, width=30,
                          bg="#d82e2e", fg="white", font=("Segoe UI", 11, "bold"), pady=6, relief="ridge", cursor="hand2")
    btn_close.pack(pady=24)

    admin_window.transient(parent)
    admin_window.grab_set()
    admin_window.mainloop()

def choose_user_type_and_login(root):
    login_window = tk.Toplevel(root)
    login_window.grab_set()
    login_window.title("Login Type")
    login_window.geometry("300x220")
    login_window.resizable(False, False)
    selected_type = tk.StringVar(value="User")
    tk.Label(login_window, text="Select User Type to Login:", font=("Segoe UI", 12, "bold")).pack(padx=20, pady=15)
    tk.Radiobutton(login_window, text="User", variable=selected_type, value="User", font=("Segoe UI", 11)).pack(pady=5)
    tk.Radiobutton(login_window, text="Admin", variable=selected_type, value="Admin", font=("Segoe UI", 11)).pack(pady=5)
    def proceed():
        choice = selected_type.get()
        login_window.destroy()
        if choice == "User":
            if user_login_check(parent=root):
                start_gui(root)
            else:
                root.quit()
        else:
            if admin_login_check(parent=root):
                admin_password_management_gui(parent=root)
                root.deiconify()
            else:
                messagebox.showerror("Access Denied", "Failed to login as Admin.", parent=root)
                # Do NOT quit; allow the user to retry or close the window themselves.
                return
    tk.Button(login_window, text="Proceed", command=proceed, width=22, bg="#0078d7", fg="white",
              activebackground="#005a9e", relief="raised", cursor="hand2").pack(pady=15)
    login_window.mainloop()

def get_start_date():
    with open(START_DATE_FILE) as file:
        return datetime.datetime.strptime(file.read().strip(), "%d-%m-%Y")

def check_expiry(parent=None):
    if datetime.datetime.today() > get_start_date() + datetime.timedelta(days=300):
        messagebox.showerror("Error", "The ID generator has expired. Please contact support.", parent=parent)
        sys.exit()

def generate_patient_id():
    prefix = 'GKNMH-CERWP-'
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    existing_ids = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
    num = 1000
    while f"{prefix}{num}" in existing_ids:
        num += 1
    return f"{prefix}{num}"

def calculate_age(dob_str, reference_str):
    try:
        dob = datetime.datetime.strptime(dob_str, "%d-%m-%Y")
        ref = datetime.datetime.strptime(reference_str, "%d-%m-%Y")
        return ref.year - dob.year - ((ref.month, ref.day) < (dob.month, dob.day))
    except:
        return ""

def validate_date(date_text):
    try:
        d = datetime.datetime.strptime(date_text, "%d-%m-%Y")
        return d <= datetime.datetime.today()
    except:
        return False

def generate_qr_code(data, qr_filename):
    qr = qrcode.QRCode(version=1, box_size=12, border=6)
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    datas = qr_img.getdata()
    new_data = [(255, 255, 255, 0) if item[:3] == (255, 255, 255) else item for item in datas]
    qr_img.putdata(new_data)
    qr_img.save(qr_filename)

def create_patient_id_card(info, qr_filename, output_filename):
    try:
        font = ImageFont.truetype("arial.ttf", 30)
        title_font = ImageFont.truetype("arial.ttf", 36)
        id_font = ImageFont.truetype("arialbd.ttf", 48)
    except:
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()
        id_font = ImageFont.load_default()
    w, h = 1240, 1748
    margin = 50
    card = Image.new("RGB", (w, h), "white")
    draw = ImageDraw.Draw(card)
    draw.rectangle([margin, margin, w - margin, h - margin], outline="black", width=5)
    if os.path.exists(LOGO_FILE):
        logo = Image.open(LOGO_FILE).resize((w - 2 * margin, 200))
        card.paste(logo, (margin, margin))
    draw.text(((w - draw.textlength("Patient ID Card", title_font)) // 2, margin + 210), 
        "Patient ID Card", font=title_font, fill="red")
    draw.text(((w - draw.textlength(info["id"], id_font)) // 2, margin + 270), info["id"], font=id_font, fill="blue")
    x_label = margin + 60
    x_colon = x_label + 300
    x_value = x_colon + 20
    y_start = margin + 400
    y_gap = 70
    details = [
        ("Patient Name", info["name"]),
        ("Date of Birth", info["dob"]),
        ("Age", f"{info['age']} years"),
        ("Gender", info["gender"]),
        ("Care Of", info["care_of"]),
        ("Phone No", info["phone"]),
        ("Registration Date", info["registration_date"])
    ]
    for idx, (label, value) in enumerate(details):
        y = y_start + idx * y_gap
        draw.text((x_label, y), label, font=font, fill="black")
        draw.text((x_colon, y), ":", font=font, fill="black")
        draw.text((x_value, y), value, font=font, fill="black")
    qr = Image.open(qr_filename).resize((400, 400))
    card.paste(qr, (w - 400 - margin, y_start), qr)
    draw.text((x_label, h - 700), "BP: ________ mm/Hg     Pulse: ______/min", font=font, fill="black")
    draw.text((x_label, h - 600), "Blood Sugar: FBS/RBS ________ mgs/dl", font=font, fill="black")
    draw.text((x_label, h - 500), "Oral:", font=font, fill="black")
    card.save(output_filename, dpi=(300, 300))

def write_to_excel(info, qr_path):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    timestamp = datetime.datetime.now().isoformat()
    sheet.append([
        info["id"], info["name"], info["dob"], info["age"], info["gender"],
        info["care_of"], info["phone"], qr_path, info["registration_date"], timestamp
    ])
    wb.save(EXCEL_FILE)

def write_to_pictures_excel(info):
    wb = openpyxl.load_workbook(PICTURES_EXCEL)
    sheet = wb.active
    timestamp = datetime.datetime.now().isoformat()
    sheet.append([
        info["id"], info["name"], info["dob"], info["age"], info["gender"],
        info["care_of"], info["phone"], info["registration_date"], timestamp
    ])
    wb.save(PICTURES_EXCEL)

def open_image_default_viewer(image_path):
    try:
        if platform.system() == "Windows":
            os.startfile(image_path)
        elif platform.system() == "Darwin":
            subprocess.call(["open", image_path])
        else:
            subprocess.call(["xdg-open", image_path])
    except Exception as e:
        print(f"Failed to open image: {e}")

def print_image_default(image_path):
    try:
        if platform.system() == "Windows":
            os.startfile(image_path, "print")
        else:
            subprocess.call(["lp", image_path])
    except Exception as e:
        print(f"Failed to print image: {e}")

def reset_form():
    global name_entry, dob_entry, gender_combobox, care_of_entry, phone_entry, calendar_widget, age_var
    if name_entry: name_entry.delete(0, tk.END); name_entry.mark_error(False)
    if dob_entry: dob_entry.delete(0, tk.END); dob_entry.mark_error(False)
    if calendar_widget: calendar_widget.selection_clear()
    if gender_combobox: gender_combobox.set(''); gender_combobox.config(background='white')
    if care_of_entry: care_of_entry.delete(0, tk.END); care_of_entry.mark_error(False)
    if phone_entry: phone_entry.delete(0, tk.END); phone_entry.mark_error(False)
    if age_var: age_var.set("")

def submit_form():
    global name_entry, dob_entry, gender_combobox, care_of_entry, phone_entry, calendar_widget, age_var
    check_expiry()
    for ctl in [name_entry, dob_entry, care_of_entry, phone_entry]: ctl.mark_error(False)
    gender_combobox.config(background="white")
    name = name_entry.get().strip()
    dob = dob_entry.get().strip()
    gender = gender_combobox.get().strip()
    care_of = care_of_entry.get().strip()
    phone = phone_entry.get().strip()
    error_found = False
    if not name: name_entry.mark_error(True); error_found = True
    if not dob or not validate_date(dob): dob_entry.mark_error(True); error_found = True
    if not gender: gender_combobox.config(background=ERROR_BORDER_COLOR); error_found = True
    if not phone.isdigit() or len(phone) != 10: phone_entry.mark_error(True); error_found = True
    if error_found:
        messagebox.showerror("Error", "Please fix the highlighted fields before submitting.")
        return
    age = calculate_age(dob, datetime.datetime.today().strftime("%d-%m-%Y"))
    patient_id = generate_patient_id()
    reg_date = datetime.datetime.today().strftime("%d-%m-%Y")
    qr_filename = os.path.join(ID_OUTPUT_DIR, f"{patient_id}_qr.png")
    generate_qr_code(patient_id, qr_filename)
    output_filename = os.path.join(ID_OUTPUT_DIR, f"{patient_id}.png")
    patient_info = {
        "id": patient_id, "name": name, "dob": dob, "age": age,
        "gender": gender, "care_of": care_of, "phone": phone,
        "registration_date": reg_date
    }
    create_patient_id_card(patient_info, qr_filename, output_filename)
    write_to_excel(patient_info, qr_filename)
    try:
        shutil.copy(output_filename, PICTURES_SUBDIR)
    except Exception as e:
        print(f"Failed copying to Pictures folder: {e}")
    write_to_pictures_excel(patient_info)
    print_image_default(output_filename)
    reset_form()
    try: os.remove(qr_filename)
    except: pass

def start_gui(root):
    global name_entry, dob_entry, gender_combobox, care_of_entry, phone_entry, calendar_widget, age_var
    app = tk.Toplevel(root)
    app.title("Patient ID Generator")
    app.geometry("1280x820")
    app.configure(background="#f8f9fa")
    app.resizable(True, True)
    # ---- Layout Frames ----
    outer = tk.Frame(app, bg="#f8f9fa")
    outer.pack(expand=True, fill="both")
    form_frame = tk.Frame(outer, bg="#f8f9fa")
    form_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nws")
    preview_block = tk.Frame(outer, bg="#f8f9fa")
    preview_block.grid(row=0, column=1, padx=10, pady=12, sticky="nesw")
    # ---- FORM FIELDS ----
    tf = tk.Frame(form_frame, bg="#f8f9fa", padx=10, pady=15)
    tf.grid(row=0, column=0)
    tk.Label(tf, text="Patient Name:", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=0, column=0, sticky="w", pady=7)
    name_entry = InteractiveEntry(tf, width=32, font=("Segoe UI", 12))
    name_entry.grid(row=0, column=1, pady=7, padx=5)
    ToolTip(name_entry, "Enter full legal name of patient")
    tk.Label(tf, text="Date of Birth (dd-mm-yyyy):", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=1, column=0, sticky="w", pady=7)
    dob_entry = InteractiveEntry(tf, width=32, font=("Segoe UI", 12))
    dob_entry.grid(row=1, column=1, pady=7, padx=5)
    ToolTip(dob_entry, "Format: dd-mm-yyyy. Date cannot be future date.")
    tk.Label(tf, text="Calendar to Set DOB:", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=2, column=0, sticky="w", pady=7)
    calendar_widget = Calendar(tf, date_pattern="dd-mm-yyyy", mindate=datetime.datetime(1900, 1, 1), maxdate=datetime.datetime.today(), background='white', foreground='black',
                                 headersbackground='#0078d7', normalbackground='white', normalforeground='black', weekendbackground='#e6f0ff', weekendforeground='black')
    calendar_widget.grid(row=2, column=1, pady=7)
    calendar_widget.selection_clear()
    tk.Label(tf, text="Age (in years):", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=3, column=0, sticky="w", pady=7)
    age_var = tk.StringVar()
    age_label = tk.Label(tf, textvariable=age_var, width=28, anchor="w", relief="sunken", font=("Segoe UI", 12), bg="white", fg="#333")
    age_label.grid(row=3, column=1, pady=7, padx=5)
    tk.Label(tf, text="Gender:", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=4, column=0, sticky="w", pady=7)
    gender_combobox = Combobox(tf, values=["Male", "Female", "Other"], state="readonly", width=28, font=("Segoe UI", 12))
    gender_combobox.grid(row=4, column=1, pady=7, padx=5)
    ToolTip(gender_combobox, "Select patient's gender")
    tk.Label(tf, text="Care Of:", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=5, column=0, sticky="w", pady=7)
    care_of_entry = InteractiveEntry(tf, width=32, font=("Segoe UI", 12))
    care_of_entry.grid(row=5, column=1, pady=7, padx=5)
    ToolTip(care_of_entry, "Name of guardian/caretaker if applicable")
    tk.Label(tf, text="Phone No:", anchor="w", width=20, font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#0078d7").grid(row=6, column=0, sticky="w", pady=7)
    phone_entry = InteractiveEntry(tf, width=32, font=("Segoe UI", 12))
    phone_entry.grid(row=6, column=1, pady=7, padx=5)
    ToolTip(phone_entry, "10-digit mobile number without country code")
    btn_generate = tk.Button(tf, text="Generate ID Card", width=32, command=submit_form,
        bg="#0078d7", fg="white", activebackground="#005a9e", relief="raised", cursor="hand2", font=("Segoe UI", 12, "bold"))
    btn_generate.grid(row=7, column=0, columnspan=2, pady=16, padx=5)
    def preview_last_id():
        latest_files = sorted([f for f in os.listdir(ID_OUTPUT_DIR) if f.endswith(".png") and not f.startswith("preview_")],
                              key=lambda x: os.path.getmtime(os.path.join(ID_OUTPUT_DIR, x)), reverse=True)
        if latest_files:
            open_image_default_viewer(os.path.join(ID_OUTPUT_DIR, latest_files[0]))
        else:
            messagebox.showwarning("No Preview", "No ID card has been generated yet.", parent=app)
    btn_preview = tk.Button(tf, text="Preview Last ID Card", width=32, command=preview_last_id,
        bg="#555", fg="white", activebackground="#333", relief="raised", cursor="hand2", font=("Segoe UI", 11))
    btn_preview.grid(row=8, column=0, columnspan=2, pady=4, padx=5)

    # --- LIVE PREVIEW SCROLLABLE ---
    preview_frame = tk.Frame(preview_block, relief="groove", bd=3, bg="white")
    preview_frame.grid(row=0, column=0, padx=8, pady=8, sticky="nsew")
    preview_label = tk.Label(preview_frame, text="Live ID Card Preview", font=("Segoe UI", 14, "bold"), fg="#b22222", bg="white")
    preview_label.pack(pady=5)

    preview_canvas_container = tk.Canvas(preview_frame, width=380, height=645, bg="white", highlightthickness=0)
    preview_canvas_container.pack(side="left", fill="both", expand=False)
    scrollbar = tk.Scrollbar(preview_frame, orient="vertical", command=preview_canvas_container.yview)
    scrollbar.pack(side="right", fill="y")
    preview_canvas_container.config(yscrollcommand=scrollbar.set)
    preview_inner_frame = tk.Frame(preview_canvas_container, bg="white")
    preview_canvas_container.create_window((0,0), window=preview_inner_frame, anchor='nw')
    preview_canvas = tk.Label(preview_inner_frame, bg="white", relief="ridge", bd=2)
    preview_canvas.pack()
    def update_scroll_region(event):
        preview_canvas_container.config(scrollregion=preview_canvas_container.bbox("all"))
    preview_inner_frame.bind("<Configure>", update_scroll_region)
    preview_enabled = tk.BooleanVar(value=True)
    def create_temp_id_card(temp_path):
        info = {
            "id": "PREVIEW-ID",
            "name": name_entry.get().strip() or ".................",
            "dob": dob_entry.get().strip() or "dd-mm-yyyy",
            "age": age_var.get() or "........",
            "gender": gender_combobox.get().strip() or "..............",
            "care_of": care_of_entry.get().strip() or "..............",
            "phone": phone_entry.get().strip() or ".............",
            "registration_date": datetime.datetime.today().strftime("%d-%m-%Y")
        }
        temp_qr = temp_path.replace(".png", "_qr.png")
        generate_qr_code(info["id"], temp_qr)
        create_patient_id_card(info, temp_qr, temp_path)
        try: os.remove(temp_qr)
        except: pass
    preview_img = [None]
    def update_preview(event=None):
        if not preview_enabled.get():
            preview_canvas.config(image="", text="Preview disabled", bg="white", font=("Segoe UI", 14, "italic"))
            preview_canvas.image = None
            return
        temp_preview_path = os.path.join(ID_OUTPUT_DIR, "preview_realtime.png")
        create_temp_id_card(temp_preview_path)
        try:
            img = Image.open(temp_preview_path)
            img_full = img.resize((380, 1290), Image.LANCZOS)
            preview_img[0] = img_full
            preview_canvas.config(image=None)
            preview_canvas.image = None
            display_preview_part(0)
        except Exception as e:
            preview_canvas.config(text=f"Preview unavailable: {e}", font=("Segoe UI", 12), bg="white")
            preview_canvas.image = None
        finally:
            try: os.remove(temp_preview_path)
            except: pass
    def display_preview_part(scroll_val):
        if preview_img[0] is None: return
        y = int(scroll_val)
        region = preview_img[0].crop((0, y, 380, y+645))
        img_tk = ImageTk.PhotoImage(region)
        preview_canvas.config(image=img_tk, text="", bg="white")
        preview_canvas.image = img_tk
    def on_mousewheel(event):
        delta = -1*(event.delta//120)
        canvas = preview_canvas_container
        canvas.yview_scroll(delta, "units")
        curr = int(canvas.canvasy(0))
        max_scroll = preview_img[0].height - 645 if preview_img[0] else 0
        val = max(0, min(curr, max_scroll))
        display_preview_part(val)
    preview_canvas_container.bind_all("<MouseWheel>", on_mousewheel)
    btn_toggle = tk.Checkbutton(preview_frame, text="Show Live ID Preview", variable=preview_enabled, bg="white", font=("Segoe UI", 11), command=update_preview)
    btn_toggle.pack(pady=(10,4))
    for widget in [name_entry, dob_entry, care_of_entry, phone_entry]:
        widget.bind("<KeyRelease>", update_preview)
        widget.bind("<FocusOut>", update_preview)
    gender_combobox.bind("<<ComboboxSelected>>", update_preview)
    dob_entry.bind("<FocusOut>", lambda e: [sync_dob_field_to_calendar(), update_preview()])
    dob_entry.bind("<KeyRelease>", lambda e: [sync_dob_field_to_calendar(), update_preview()])
    calendar_widget.bind("<<CalendarSelected>>", lambda e: [sync_calendar_to_dob_field(), update_preview()])
    def sync_dob_field_to_calendar(event=None):
        entered_dob = dob_entry.get().strip()
        if validate_date(entered_dob):
            age = calculate_age(entered_dob, datetime.datetime.today().strftime("%d-%m-%Y"))
            age_var.set(str(age))
            try:
                d = datetime.datetime.strptime(entered_dob, "%d-%m-%Y")
                calendar_widget.selection_set(d)
            except ValueError: pass
        else: age_var.set("")
    def sync_calendar_to_dob_field(event=None):
        selected_date = calendar_widget.get_date()
        dob_entry.delete(0, tk.END)
        dob_entry.insert(0, selected_date)
        if validate_date(selected_date):
            age = calculate_age(selected_date, datetime.datetime.today().strftime("%d-%m-%Y"))
            age_var.set(str(age))
        else:
            age_var.set("")
    sync_dob_field_to_calendar()
    update_preview()
    outer.grid_columnconfigure(0, weight=1, minsize=420)
    outer.grid_columnconfigure(1, weight=1, minsize=420)
    outer.grid_rowconfigure(0, weight=1)
    app.mainloop()

if __name__ == "__main__":
    setup_dirs_and_files()
    root = tk.Tk()
    root.withdraw()
    choose_user_type_and_login(root)
    root.mainloop()
